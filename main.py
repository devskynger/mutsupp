import pandas as pd
#перейди на последнюю строку говнюк
import vk_api, threading, traceback, openpyxl, random, os, time
from vk_api.longpoll import VkLongPoll, VkEventType
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill
from vk_api.upload import VkUpload
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType

ServiceToken1 = 'b115d9a5b115d9a5b115d9a5bfb207e21bbb115b115d9a5d2d63d72eca97b73a59da34c'
ServiceToken2 = 'e51f6cc3e51f6cc3e51f6cc313e5630d51ee51fe51f6cc38776053b4afe81fc29de2890'
ServiceToken3 = '21f4181621f4181621f418166c22e5d00e221f421f41816426457688139458300c796fe'
ServiceToken4 = '8c4d16958c4d16958c4d1695bb8f5cdd8d88c4d8c4d1695efdc2d5520978aa39eb59ed6'
ServiceToken5 = 'a9b0c790a9b0c790a9b0c7905daaa10c89aa9b0a9b0c790ca21fc607f75bb761717d594'
ServiceToken6 = 'b89e9959b89e9959b89e995912bb8f5243bb89eb89e9959db0fa5467fb47fae87d507bb'
sessionSrv1 = vk_api.VkApi(token=ServiceToken1)
vkSrv1 = sessionSrv1.get_api()
sessionSrv2 = vk_api.VkApi(token=ServiceToken2)
vkSrv2 = sessionSrv2.get_api()
sessionSrv3 = vk_api.VkApi(token=ServiceToken3)
vkSrv3 = sessionSrv3.get_api()
sessionSrv4 = vk_api.VkApi(token=ServiceToken4)
vkSrv4 = sessionSrv3.get_api()
sessionSrv5 = vk_api.VkApi(token=ServiceToken5)
vkSrv5 = sessionSrv3.get_api()
sessionSrv6 = vk_api.VkApi(token=ServiceToken6)
vkSrv6 = sessionSrv3.get_api()

session = vk_api.VkApi(token="vk1.a.auoBmmYDE3e8VMGSEtC50u_Zw_qdqgsawMu67QrNhTcQMUUvx70YR4EiPSwcZj-bxj9VBA0JoBdPRNTMgjlKy8yCTKuNVyZtq7B4ZB3yPR7A-J9NPi549ULYiSL_RNDugdZaeff1k9ugRG3TRURGs_tb8RBzOSEajZ7GiFeQL7dz_5cjfVGkm3FcJcmFNAn2c2I4pvCcQei6J-C-scK46g")
vk = session.get_api()
longpoll = VkLongPoll(session)

ses_gr = vk_api.VkApi(token='vk1.a.rn1i54NNIIHkUHQTihMNZ83oyd7Qy0mHD-OhgDNDfgJYLPssmMIx6kCYlWvUHHm08VBELne98qib2gpYf-lzrUEZF3rq-JqwP-Y2LA70ZFikgW5cGGvyaUxSfJ20RBxUFS-oaISt6HpA-yJk2M7QR6WU847SlpL-xrdxn3B8Irg__p8gVf_2ayR4Pbm7-Ado3nQQtvwgljG6EjfTsJoRyw')
vk_gr = ses_gr.get_api()
lngl_gr = VkBotLongPoll(ses_gr, 222006388)
prEvent = None
ses_gr1 = vk_api.VkApi(token='vk1.a.F7LL-20u-GRXdWigtve9EHlJPZ-j3mAtF8NG6pBoHOmS5KtXhWt3cZqAM_IZdsrvYkXZu1t2TAc6Se3Yc9GV6m9vAcBdqOV80ErS6mBT0VgMn7aJY4xUQriAVn26mUd_uhBH7TnzE8cAvpy1HxSsaDs7N90CfbtlXGOB0XiO3UlvtLa6pZKWSbWtLcq2BEBTxu8GJW6QzUhyKda0thQuMA')
vk_gr1 = ses_gr1.get_api()


msg_send_all_posts = {
	
	'premium': '• Выполняем задания:\n 1) Лайк \n2) Полноценный комментарий от трех слов ( комментарии со смайликами, слова по типу: «топ», «вау», «круто», «класс», а также комментарии не соответствующие тематике поста-запрещены!!! Те, кто будет писать такие комментарии, будут ИСКЛЮЧЕНЫ)',
	'standart': '• Выполняем задания:\n 1) Лайк \n2) Комментарий',
	'like': '• Выполняем задания:\n 1) Лайк',
	'time': '• Выполняем задания:\n 1) Лайк \n2) Полноценный комментарий от трех слов ( комментарии со смайликами, слова по типу: «топ», «вау», «круто», «класс», а также комментарии не соответствующие тематике поста-запрещены!!! Те, кто будет писать такие комментарии, будут ИСКЛЮЧЕНЫ)',

}


def main(): # Запуск бота
	global prEvent
	try:
		for event in lngl_gr.listen():
			if event.obj != prEvent: 
				prEvent = event.obj
				threading.Thread(target=EventHandler, args=(event, )).start()

	except BaseException as e:
		print(e)
		threading.Thread(target=main).start() 
	print('отключаюсь main')


def EventHandler(event):
	try:
		if event.type == VkBotEventType.MESSAGE_NEW: # Сообщения   


			if event.from_chat:
				id = event.obj['message']['peer_id']-2000000000
				peer_id = event.obj['message']['peer_id']
				text = event.obj['message']['text']
				userid = event.obj['message']['from_id']

				if id != 1:
					bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
					if text == "!бот" and (userid == 478875154 or userid == 730993538 or userid == 251759221):
						if str(peer_id) not in bot_in_chats:
							open(f"data/bot_in_chats.txt", 'a', encoding="utf-8").write(str(peer_id)+'\n')
							vk_gr.messages.send(peer_id=peer_id, message='Данный чат добавлен в БД\n\n*не забудьте выдать админку!', random_id=0)
						else:
							vk_gr.messages.send(peer_id=peer_id, message='На месте!', random_id=0)

				else:
					# Отправка постов
					if 'во все ' in text:
						try:
							post = text.replace('во все  ','').split()[0]
							if 'https://' in post: 
								if "vk" not in post: # не пост вк
									vk_gr.messages.send(user_id=id, message='⛔ Сюда можно кидать только ссылки на ПОСТЫ ВК!', random_id=0)
								else:
									post = PostHandler(post)
									post = post.replace(' ','')
									try:
										ForOwnANdItem = post.replace("https://vk.com/wall", "").split("_")
										int(ForOwnANdItem[0])
										int(ForOwnANdItem[1])
									except BaseException:
										vk_gr.messages.send(user_id=id, message='⛔ Не правильная ссылка! (киньте стандартную ссылку типа: "https://vk.com/wall-123456789_123" или обратись в @id478875154(поддержку))', random_id=0)
										threading.Thread(target=event_handler_group).start()
										return

									partners_posts = open('data/partners_posts.txt', encoding="utf-8").read() if os.path.exists('data/partners_posts.txt') else ''
									if post in partners_posts:
										vk_gr.messages.send(chat_id=id, message=f'⛔ Этот пост уже записан!\n{post}', random_id=0)
									else:
										open('data/partners_posts.txt', 'a', encoding='utf-8').write(post+'\n')
										vk_gr.messages.send(chat_id=id, message=f'✅ Принял\n{post}', random_id=0)
						except BaseException as e:
							vk_gr.messages.send(chat_id=id, message=f'Ошибка: {e}', random_id=0)



			if event.from_user: 
				id = event.obj['message']['from_id']
				mid = event.obj['message']['id']
				text = event.obj['message']['text'].lower()
				weekday = datetime.today().weekday()

				if text == '!бот':
					vk_gr.messages.send(user_id=id, message='На месте!', random_id=0)


				try:
					vk_gr.messages.send(user_id=id, message='🔎 Получаю информацию о вас...', random_id=0)
					info = info_chats(id)
				except BaseException:
					vk_gr.messages.send(user_id=id, message='Ошибка со стороны сервера. Попробуйте ещё раз или обратитесь в @id478875154(поддержку)', random_id=0)
					return

				group_name = info['group_name']
				group_num = info['group_num']
				in_chats = info['in_chats'] 
				is_admin = info['is_admin']
				in_like = info['in_like']
				like_group_num = info['like_group_num']

				if in_chats == False:
					vk_gr.messages.send(user_id=id, message='🤨⛔ Вас нет ни в одном из чатов. Попроситее её добавить вас: https://vk.com/id478875154', random_id=0)
				else:


					all_users = open(f"data/all_users.txt", encoding="utf-8").read().split() if os.path.exists(f"data/all_users.txt") else []
					if str(id) not in all_users:
						vk_gr.messages.send(user_id=id, message= 
							"""

								Привествую, это системный бот проекта по продвижению EasyStart Promotion.\n\n

								Если вы зашли к нам в первый раз и еще не были ознакомлены с регламентом проекта, 
								просьба написать «+» в сообщения официального сообщества проекта https://vk.com/1easystart\n\n

								Наши специалисты в течение дня с вами свяжутся, ознакомят вас с правилами и ответят на все ваши вопросы!\n\n

								Теперь внимательно прочитайте эти пункты:\n
								- Если вы будете комментировать посты от имени своей группы, то пришлите следующее сообщение:
								группа коммы ССЫЛКА_НА_ГРУППУ
								\n\n
								С уважением, команда EasyStart Team❤

							""",
							random_id=0
							)
						open(f"data/all_users.txt", 'a', encoding="utf-8").write(str(id)+'\n')


					if 'группа ' in text and len(text.split()) == 3:
						if 'коммы ' in text:
							try:
								group = text.replace("группа коммы ", "")
								if '[' in group and ']' in group:
									group = 'https://vk.com/'+group.replace('[', '').replace(']','').split('|')[0]
								group = group.replace('https://vk.com/','')
									
								group = str(vk_gr.groups.getById(group_id=group)[0]['id'])
								open(f'data/{id}group_for_comms.txt', 'w', encoding='utf-8').write(str(group))
								vk_gr.messages.send(user_id=id, message='Записал группу, с которой вы будете оставлять комментарии. Если на каком-то из постов комментарии от группы оставлять нельзя, то оставляйте их от имени своей страницы, бот будет засчитывать такие комментарии', random_id=0)
							except BaseException:
								print(traceback.format_exc())
								vk_gr.messages.send(user_id=id, message='Ошибка', random_id=0)
						elif 'посты ' in text:
							pass 


					# Приём постов
					if 'https://' in text and 'группа ' not in text: 

						if 'лайк ' in text and len(text.split())== 2:
							text = text.replace('лайк ', '')
							if in_like == False:
								vk_gr.messages.send(user_id=id, message='⛔ Вас нет в группе Like!', random_id=0)
							else:
								if "vk" not in text: # не пост вк
									vk_gr.messages.send(user_id=id, message='⛔ Сюда можно кидать только ссылки на ПОСТЫ ВК!', random_id=0)
								else:
									text = PostHandler(text)
									text = text.replace(' ','')
									try:
										ForOwnANdItem = text.replace("https://vk.com/wall", "").split("_")
										int(ForOwnANdItem[0])
										int(ForOwnANdItem[1])
									except BaseException:
										vk_gr.messages.send(user_id=id, message='⛔ Не правильная ссылка! (киньте стандартную ссылку типа: "https://vk.com/wall-123456789_123" или обратись в @id478875154(поддержку))', random_id=0)
										threading.Thread(target=event_handler_group).start()
										return

									g_nums = []
									g_nums.append(int(like_group_num))
									ev = is_even(int(like_group_num)) 
									if ev == True:
										g_nums.append(int(like_group_num)-1)
									else:
										g_nums.append(int(like_group_num)+1)

									if g_nums[0] > g_nums[1]:
										users_send_post = open(f"data/people_like_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/people_like_{str(g_nums[0])}_{str(g_nums[1])}.txt") else []
										posts = open(f"data/posts_like_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/posts_like_{str(g_nums[0])}_{str(g_nums[1])}.txt") else []
									else:
										users_send_post = open(f"data/people_like_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/people_like_{str(g_nums[1])}_{str(g_nums[0])}.txt") else []
										posts = open(f"data/posts_like_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/posts_like_{str(g_nums[1])}_{str(g_nums[0])}.txt") else []

									prom_group = promotion_group("like",like_group_num) # Какая группа скидывает посты сегодня 

									if str(prom_group) == str(like_group_num):
										if str(id) not in users_send_post: 
											if text not in posts:

												
												if g_nums[0] > g_nums[1]:
													open(f"data/posts_like_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'a', encoding="utf-8").write(str(text)+'\n')
													open(f"data/people_like_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'a', encoding="utf-8").write(str(id)+'\n')
												else:
													open(f"data/posts_like_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'a', encoding="utf-8").write(str(text)+'\n')
													open(f"data/people_like_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'a', encoding="utf-8").write(str(id)+'\n')
												
												vk_gr.messages.send(user_id=id, message='✅ Ваша ссылка принята! Она будет опубликована в рабочей беседе (группе) Like в 19:00 по МСК', random_id=0)
												vk_gr.messages.send(user_id=id, message=f'Сейчас вы находитесь в группе {like_group_num} like. Если информация не верна, то обратитесь в @id478875154(поддержку)', random_id=0)
											else:
												vk_gr.messages.send(user_id=id, message='⛔ Этот пост уже кидали сегодня!', random_id=0)
										else:
											vk_gr.messages.send(user_id=id, message='⛔ Вы сегодня уже кидали пост!', random_id=0)
									else:
										vk_gr.messages.send(user_id=id, message=f'⛔ Сегодня скидывает посты {prom_group} like группа!', random_id=0)


						else: 
							tms = datetime.now().strftime('%H')
							time_tms = ['00','01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16']
							if group_name != "time" or (group_name == "time" and (weekday == 5 or weekday == 6) and tms in time_tms):
								if "vk" not in text: # не пост вк
									vk_gr.messages.send(user_id=id, message='⛔ Сюда можно кидать только ссылки на ПОСТЫ ВК!', random_id=0)
								else:
									text = PostHandler(text)
									text = text.replace(' ','')
									try:
										ForOwnANdItem = text.replace("https://vk.com/wall", "").split("_")
										int(ForOwnANdItem[0])
										int(ForOwnANdItem[1])
									except BaseException:
										vk_gr.messages.send(user_id=id, message='⛔ Не правильная ссылка! (киньте стандартную ссылку типа: "https://vk.com/wall-123456789_123" или обратись в @id478875154(поддержку))', random_id=0)
										return

									g_nums = []
									g_nums.append(int(group_num))
									ev = is_even(int(group_num)) 
									if ev == True:
										g_nums.append(int(group_num)-1)
									else:
										g_nums.append(int(group_num)+1)

									if g_nums[0] > g_nums[1]:
										users_send_post = open(f"data/people_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/people_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else []
										posts = open(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else []
									else:
										users_send_post = open(f"data/people_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/people_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else []
										posts = open(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else []

									prom_group = promotion_group(group_name,group_num) # Какая группа скидывает посты сегодня 

									if str(prom_group) == str(group_num):
										if str(id) not in users_send_post: 
											if text not in posts:

												
												if g_nums[0] > g_nums[1]:
													open(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'a', encoding="utf-8").write(str(text)+'\n')
													open(f"data/people_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'a', encoding="utf-8").write(str(id)+'\n')
												else:
													open(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'a', encoding="utf-8").write(str(text)+'\n')
													open(f"data/people_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'a', encoding="utf-8").write(str(id)+'\n')
												
												vk_gr.messages.send(user_id=id, message=f'✅ Ваша ссылка принята в группу {group_num} {group_name}! Она будет опубликована в рабочей беседе (группе) в 17:00 по МСК', random_id=0)
												if group_name != 'like' and in_like == True:
													vk_gr.messages.send(user_id=id, message='* Если вы хотите скинуть пост в группу Like, то напишите следующее сообщение:\n\nлайк ЗДЕСЬ_ВАШ_ПОСТ', random_id=0)
											else:
												vk_gr.messages.send(user_id=id, message='⛔ Этот пост уже кидали сегодня!', random_id=0)
										else:
											vk_gr.messages.send(user_id=id, message='⛔ Вы сегодня уже кидали пост!', random_id=0)
									else:
										vk_gr.messages.send(user_id=id, message=f'⛔ Сегодня скидывает посты {prom_group} {group_name} группа!', random_id=0)
							else:
								vk_gr.messages.send(user_id=id, message=f'⛔ Группа Time скидывает посты только в Субботу и Воскресенье c 00:00 до 17:00!', random_id=0)


	except BaseException:
		print(traceback.format_exc())


def promotion_group(g_name,g_num):

	g_nums = []
	g_nums.append(int(g_num))
	ev = is_even(int(g_num)) 
	if ev == True:
		g_nums.append(int(g_num)-1)
	else:
		g_nums.append(int(g_num)+1)

	if g_nums[0] > g_nums[1]:
		prom_gr = open(f"data/prom_gr_{g_name}_{g_nums[0]}_{g_nums[1]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{g_name}_{g_nums[0]}_{g_nums[1]}.txt") else None
	else:
		prom_gr = open(f"data/prom_gr_{g_name}_{g_nums[1]}_{g_nums[0]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{g_name}_{g_nums[1]}_{g_nums[0]}.txt") else None
	return prom_gr


def info_chats(id):
	is_adm = False
	in_like = False
	like_group_num = None
	allUsers = []
	RTRN = {'in_chats': int(id) in allUsers, 'group_name': None, "group_num": None, "is_admin": None, "in_like": None}
	bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
	for p_id in bot_in_chats:
		inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
		chat_name = inf['items'][0]['chat_settings']['title'].lower()
		if 'easystart' in chat_name and 'группа' in chat_name:
			members = vk_gr.messages.getConversationMembers(peer_id=p_id)
			admins = [member['member_id'] for member in members['items'] if member.get('is_admin')]

			if is_adm == False:
				is_adm = int(id) in admins
			for i in members['items']:
				allUsers.append(i['member_id'])
				if id == i['member_id']:
					if RTRN['group_name'] == None or (RTRN['group_name'] == 'like' and chat_name.split()[1] != 'like'):
						RTRN =  {'in_chats': True, 'group_name': chat_name.split()[1], "group_num": chat_name.split()[0], "is_admin": is_adm}
					if chat_name.split()[1] == "like":
						in_like = True
						like_group_num = chat_name.split()[0]
	RTRN['in_like'] = in_like		
	RTRN['like_group_num'] = like_group_num			
	return RTRN


def timer_check():
	time_check_people = ['00:00:00','01:00:00','02:00:00','03:00:00','04:00:00','05:00:00','06:00:00','07:00:00','08:00:00',
	'09:00:00','10:00:00','11:00:00','12:00:00','13:00:00','14:00:00','15:00:00','16:00:00','18:00:00','19:00:00','20:00:00',
	'21:00:00','22:00:00','23:00:00']
	while True:
		try:
			time_ch = datetime.now().strftime('%H:%M:%S')

			# Отправка напоминания о тос что публикуется группа или нет
			if time_ch == '10:00:00' or time_ch == '15:00:00':
				weekday = datetime.today().weekday()
				all_names_groups = ['premium', 'standart', 'like', 'time']
				for group_name in all_names_groups:
					if group_name != "time" or (group_name == "time" and (weekday == 5 or weekday == 6)):
						for p_id in bot_in_chats:
							inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
							chat_name = inf['items'][0]['chat_settings']['title'].lower()
							if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:
								g_num = chat_name.split()[0]
								g_nums = []
								g_nums.append(int(g_num))
								ev = is_even(int(g_num)) 
								if ev == True:
									g_nums.append(int(g_num)-1)
								else:
									g_nums.append(int(g_num)+1)

								if g_nums[0] > g_nums[1]:
									prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt") else None
								else:
									prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt") else None

								if send == True:
									if g_num == prom_gr: # у них принимаются посты
										vk_gr.messages.send(peer_id=p_id, message=f"@all Дорогие участники, жду вашу ссылку на пост, который вы хотите продвинуть, в ЛС сообщества, до 17:00 по МСК (кинуть ссылку нужно без номера группы, без отчетов о выполнении, просто ссылку)\nЛС сообщества - https://vk.com/im?sel=-222006388 ", random_id=0)
									else: # у них не принимаются
										if time_ch != '15:00:00':
											vk_gr.messages.send(peer_id=p_id, message=f"@all Дорогие участники, ваша группа сегодня не публикуется, но задания выполнить нужно ВСЕМ ОБЯЗАТЕЛЬНО! ", random_id=0)
							

			if time_ch == '00:00:00':
				weekday = datetime.today().weekday()
				if weekday == 5:
					write_prom_group(True, 'time')
				elif weekday == 6:
					write_prom_group(True, 'time')

			
			if time_ch == "17:00:00":
				weekday = datetime.today().weekday()
				try:
					check()
				except BaseException as e:
					print(traceback.format_exc())
					vk_gr.messages.send(chat_id=1, message='ошибка в проверке:\n\n'+str(e), random_id=0)
				if weekday == 5 or weekday == 6 or weekday == 0:
					if weekday == 0:
						check(g_n='time')
					else:
						check(g_n='time', ch=False)
						group_name = 'time'
						# отправка всех постов данному отделу и его группам
						bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
						for p_id in bot_in_chats:
							inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
							chat_name = inf['items'][0]['chat_settings']['title'].lower()
							if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:
								g_num = chat_name.split()[0]
								g_nums = []
								g_nums.append(int(g_num))
								ev = is_even(int(g_num)) 
								if ev == True:
									g_nums.append(int(g_num)-1)
								else:
									g_nums.append(int(g_num)+1)

								pr_gr = promotion_group(group_name, g_num)
								today_date = datetime.now().strftime("%d.%m.%y")
								partners_posts = open('data/partners_posts.txt', encoding="utf-8").read() if os.path.exists('data/partners_posts.txt') else ''
								if g_nums[0] > g_nums[1]:
									td_posts = open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read() if os.path.exists(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else ''
									td_posts += partners_posts
									msg = ''
									n = 0
									for ppst in td_posts.split():
										n+=1
										msg += f'{n}) {ppst}\n\n'
									if len(td_posts.split()) > 0:
										vk_gr.messages.send(peer_id=p_id, message="@all Посты "+str(pr_gr)+" группы\n\n"+msg_send_all_posts[group_name]+"\n\n"+partners_posts+msg, random_id=0)
								else:
									td_posts = open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read() if os.path.exists(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else ''
									td_posts += partners_posts
									msg = ''
									n = 0
									for ppst in td_posts.split():
										n+=1
										msg += f'{n}) {ppst}\n\n'
									if len(td_posts.split()) > 0:
										vk_gr.messages.send(peer_id=p_id, message="@all Посты "+str(pr_gr)+" группы\n\n"+msg_send_all_posts[group_name]+"\n\n"+partners_posts+msg, random_id=0)
						if weekday == 6:
							check(g_n='time')

				open('data/partners_posts.txt', 'w', encoding='utf-8').write('')


			if time_ch in time_check_people: # проверка на то, чтобы человек был только в одной беседе 
				threading.Thread(target=remove_duplicate_members).start()
				time.sleep(10)
				
		except BaseException:
			print(traceback.format_exc())


def remove_duplicate_members(is_like=False):
	# Сначала без лайк групп
	bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
	chat_members = {}
	for p_ids in chunks(bot_in_chats, 25):
		try:
			code = 'return [' + ','.join(f'{{"conversations": API.messages.getConversationsById({{"peer_ids":{p_id}}}), "members": API.messages.getConversationMembers({{"peer_id":{p_id}}}), "p_id": {p_id}}}' for p_id in p_ids) + '];'
			results = vk_gr1.execute(code=code)
			for result in results:
				p_id = result['p_id']
				inf = result['conversations']
				members = result['members']
				chat_name = inf['items'][0]['chat_settings']['title'].lower()

				if is_like == False:
					if 'easystart' in chat_name and 'группа' in chat_name and 'like' not in chat_name:
						admins = [member['member_id'] for member in members['items'] if member.get('is_admin')]
						for i in members['items']:
							id = i['member_id']
							if id not in admins:
								if id not in chat_members:
									chat_members[id] = [p_id]
								else:
									chat_members[id].append(p_id)
									code1='''
									var chat_id = parseInt({chat_id}) - 2000000000;
									var member_id = parseInt({member_id});
									API.messages.removeChatUser({ 'chat_id': chat_id, 'member_id': member_id });
									API.messages.removeChatUser({ 'chat_id': parseInt({p_id}), 'member_id': member_id });
									'''.format(chat_id=chat_members[id][0], member_id=id, p_id=int(p_id)-2000000000)
									vk_gr1.execute(code=code1)
				else:
					if 'easystart' in chat_name and 'группа' in chat_name and 'like' in chat_name:
						admins = [member['member_id'] for member in members['items'] if member.get('is_admin')]
						for i in members['items']:
							id = i['member_id']
							if id not in admins:
								if id not in chat_members:
									chat_members[id] = [p_id]
								else:
									chat_members[id].append(p_id)
									code1='''
									var chat_id = parseInt({chat_id}) - 2000000000;
									var member_id = parseInt({member_id});
									API.messages.removeChatUser({ 'chat_id': chat_id, 'member_id': member_id });
									API.messages.removeChatUser({ 'chat_id': parseInt({p_id}), 'member_id': member_id });
									'''.format(chat_id=chat_members[id][0], member_id=id, p_id=int(p_id)-2000000000)
									vk_gr1.execute(code=code1)
	
		except BaseException:
			print('ошибка в remove_duplicate_members:\n\n'+traceback.format_exc())
	if is_like == False:
		remove_duplicate_members(is_like=True)


def chunks(lst, n):
	"""Yield successive n-sized chunks from lst."""
	for i in range(0, len(lst), n):
		yield lst[i:i + n]


def write_prom_group(send, gr_n=None):

	if gr_n == None:
		all_names_groups = ['premium', 'standart', 'like']
	else:
		all_names_groups = [gr_n]
	RTRN = {}
	# пробегаемся по всем группам (названия)
	for group_name in all_names_groups:
		RTRN[group_name] = {}
		all_nums_groups = [] # все номера групп list
		chck_nms_gr = [] # группы которые уже собраны
		
		# собираем номера всех групп в переменную выше
		bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
		for p_id in bot_in_chats:
			inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
			chat_name = inf['items'][0]['chat_settings']['title'].lower()
			if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:
				g_num = chat_name.split()[0]
				if int(g_num) not in chck_nms_gr:
					g_nums = []
					g_nums.append(int(g_num))
					chck_nms_gr.append(int(g_num))
					ev = is_even(int(g_num)) 
					if ev == True:
						g_nums.append(int(g_num)-1)
						chck_nms_gr.append(int(g_num)-1)
					else:
						g_nums.append(int(g_num)+1)
						chck_nms_gr.append(int(g_num)+1)
					all_nums_groups.append(g_nums)

		# пробегаемся по номерам данной группы
		for g_nums in all_nums_groups:

			if g_nums[0] > g_nums[1]:
				prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt") else None
				if prom_gr == None:
					open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", 'w', encoding="utf-8").write(str(g_nums[1]))
					RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[1])
				else:
					if prom_gr == str(g_nums[0]):
						open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", 'w', encoding="utf-8").write(str(g_nums[1]))
						RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[1])
					elif prom_gr == str(g_nums[1]):
						open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", 'w', encoding="utf-8").write(str(g_nums[0]))
						RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[0])

			else:
				prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt") else None
				if prom_gr == None:
					open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", 'w', encoding="utf-8").write(str(g_nums[0]))
					RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[0])
				else:
					if prom_gr == str(g_nums[0]):
						open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", 'w', encoding="utf-8").write(str(g_nums[1]))
						RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[1])
					elif prom_gr == str(g_nums[1]):
						open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", 'w', encoding="utf-8").write(str(g_nums[0]))
						RTRN[group_name][str(g_nums[0])+"_"+str(g_nums[1])] = str(g_nums[0])

		for p_id in bot_in_chats:
			inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
			chat_name = inf['items'][0]['chat_settings']['title'].lower()
			if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:
				g_num = chat_name.split()[0]
				g_nums = []
				g_nums.append(int(g_num))
				ev = is_even(int(g_num)) 
				if ev == True:
					g_nums.append(int(g_num)-1)
				else:
					g_nums.append(int(g_num)+1)

				if g_nums[0] > g_nums[1]:
					prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[0]}_{g_nums[1]}.txt") else None
				else:
					prom_gr = open(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt", encoding="utf-8").read() if os.path.exists(f"data/prom_gr_{group_name}_{g_nums[1]}_{g_nums[0]}.txt") else None

				if send == True:
					if g_num == prom_gr: # у них принимаются посты
						vk_gr.messages.send(peer_id=p_id, message=f"@all Дорогие участники, жду вашу ссылку на пост, который вы хотите продвинуть, в ЛС сообщества, до 17:00 по МСК (кинуть ссылку нужно без номера группы, без отчетов о выполнении, просто ссылку)\nЛС сообщества - https://vk.com/im?sel=-222006388 ", random_id=0)
					else: # у них не принимаются
						vk_gr.messages.send(peer_id=p_id, message=f"@all Дорогие участники, ваша группа сегодня не публикуется, но задания выполнить нужно ВСЕМ ОБЯЗАТЕЛЬНО! ", random_id=0)
				
	return RTRN


def check(g_n=None, ch=True):
	if g_n == None:
		all_names_groups = ['premium', 'standart', 'like']
	else:
		all_names_groups = [g_n]

	# пробегаемся по всем группам (названия)
	for group_name in all_names_groups:
		all_nums_groups = [] # все номера групп list
		chck_nms_gr = [] # группы которые уже собраны

		# собираем номера всех групп в переменную выше
		bot_in_chats = open(f"data/bot_in_chats.txt", encoding="utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
		for p_id in bot_in_chats:
			inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
			chat_name = inf['items'][0]['chat_settings']['title'].lower()
			if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:

				g_num = chat_name.split()[0]
				if int(g_num) not in chck_nms_gr:
					g_nums = []
					g_nums.append(int(g_num))
					chck_nms_gr.append(int(g_num))
					ev = is_even(int(g_num)) 
					if ev == True:
						g_nums.append(int(g_num)-1)
						chck_nms_gr.append(int(g_num)-1)
					else:
						g_nums.append(int(g_num)+1)
						chck_nms_gr.append(int(g_num)+1)
					all_nums_groups.append(g_nums)


		# пробегаемся по номерам данной группы
		for g_nums in all_nums_groups:

			# записываем посты в переменную посты а также удаляем посты 
			last_date = (datetime.now()-timedelta(days=1)).strftime("%d.%m.%y")
			today_date = datetime.now().strftime("%d.%m.%y")
			posts = []
			if g_nums[0] > g_nums[1]:
				posts = open(f"data/{last_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/{last_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else []
				td_posts = open(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read() if os.path.exists(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else ''
				open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'w',encoding="utf-8").write(td_posts)
				open(f"data/posts_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'w',encoding="utf-8").write('')
				open(f"data/people_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", 'w', encoding="utf-8").write('')
			else:
				posts = open(f"data/{last_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read().split() if os.path.exists(f"data/{last_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else []
				td_posts = open(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read() if os.path.exists(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else ''
				open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'w',encoding="utf-8").write(td_posts)
				open(f"data/posts_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'w',encoding="utf-8").write('')
				open(f"data/people_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", 'w', encoding="utf-8").write('')
			psts = posts
			posts = []
			for i in psts:
				if i not in posts:
					posts.append(i)
			

			if ch == True:

				if len(posts) == 0:
					vk_gr.messages.send(chat_id=1, message=f'Проверка в группе {str(g_nums[0])}, {str(g_nums[1])} {group_name} не завершена: нет постов', random_id=0)
				else:

					chts_ids = {}

					bot_in_chats = open(f"data/bot_in_chats.txt", encoding= "utf-8").read().split() if os.path.exists(f"data/bot_in_chats.txt") else []
					for p_id in bot_in_chats:
						inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
						chat_name = inf['items'][0]['chat_settings']['title'].lower()
						if 'easystart' in chat_name and 'группа' in chat_name:
							if group_name == chat_name.split()[1]:
								if int(chat_name.split()[0]) in g_nums:
									chts_ids[chat_name.split()[0]+' '+chat_name.split()[1]] = p_id

					file_name = 'data/'+(datetime.now()-timedelta(days=1)).strftime('%d.%m.%y')+group_name+'_'+str(g_nums[0])+'_'+str(g_nums[1])+'.xlsx'

					df = pd.DataFrame()
					df.to_excel(file_name, index=False)		

					сheck_LkUsrs = CheckLikes(posts, random.choice([vkSrv1,vkSrv2,vkSrv3,vkSrv4,vkSrv5,vkSrv6]))
					check_CmUsrs = CheckComments(posts, random.choice([vkSrv1,vkSrv2,vkSrv3,vkSrv4,vkSrv5,vkSrv6]))		

					for cc in chts_ids:
						ch_id = chts_ids[cc]
						cc = cc.split()[0]
						gr_num = cc

						# Получение информации о участниках беседы 
						if g_nums[0] > g_nums[1]:
							people_file = f'data/people_for_check_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}_{ch_id}.txt'
						else:
							people_file = f'data/people_for_check_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}_{ch_id}.txt'
						
						allUsers = []
						if os.path.exists(people_file): # если файл есть, то берем людей из файла 
							allUsers = open(people_file, encoding='utf-8').read().split()

							members = vk_gr.messages.getConversationMembers(peer_id=ch_id, fields='is_admin')
							# Получение администратоов беседы
							admins = []
							for member in members['items']:
								try:
									if member['is_admin']:
										admins.append(member['member_id'])
								except KeyError:
									pass

							# Получение всех участников без админов
							allUsers1 = []
							for i in members['items']:
								id = i['member_id']
								if id not in admins:
									allUsers1.append(id)

							for i in allUsers:
								if int(i) not in allUsers1:
									allUsers.remove(i)

							open(people_file, 'w', encoding='utf-8').write('')
							for i in allUsers1:
								open(people_file, 'a', encoding='utf-8').write(str(i)+'\n')
						else: # Если нет то из беседы
							members = vk_gr.messages.getConversationMembers(peer_id=ch_id, fields='is_admin')

							# Получение администратоов беседы
							admins = []
							for member in members['items']:
								try:
									if member['is_admin']:
										admins.append(member['member_id'])
								except KeyError:
									pass

							# Получение всех участников без админов
							allUsers = []
							for i in members['items']:
								id = i['member_id']
								if id not in admins:
									allUsers.append(str(id))
							open(people_file, 'w', encoding='utf-8').write('')
							for i in allUsers:
								open(people_file, 'a', encoding='utf-8').write(str(i)+'\n')

						nms = []
						alUsExc = []
						for i in range(0, len(allUsers), 100):
							users = random.choice([vkSrv1,vkSrv2,vkSrv3,vkSrv4,vkSrv5,vkSrv6]).users.get(user_ids=allUsers[i:i+100], lang='ru')
							for user in users:
								nms.append(user['first_name'] + ' ' + user['last_name'])
								alUsExc.append('@id' + str(user['id']))

						grps_comms = []
						for i in allUsers:
							gr = ''
							try:
								gr = open(f'data/{str(i)}group_for_comms.txt', encoding='utf-8').read()
							except BaseException:
								pass
							if gr == '':
								grps_comms.append(' ')
							else:
								grps_comms.append(f'@club{gr}')
						df = pd.DataFrame({'name': nms, 'id': alUsExc, 'group':grps_comms})
						
						writer = pd.ExcelWriter(file_name, engine='openpyxl',  mode='a')

						df.to_excel(writer, sheet_name=str(gr_num), index=False)

						workbook = writer.book
						worksheet = writer.sheets[str(gr_num)]

						worksheet.column_dimensions['A'].width = 25
						worksheet.column_dimensions['B'].width = 15	
						worksheet.column_dimensions['C'].width = 15
						worksheet.column_dimensions['D'].width = 100

						writer.close()

						wb = openpyxl.load_workbook(file_name)
						ws = wb[str(gr_num)]
						
						n = 1
						for i in allUsers:
							n += 1

							
							# Проверка лайков и комментариев
							checkLiked = None
							CheckCmts = None

							notLikedPsts = []
							for kk in сheck_LkUsrs:
								if int(i) not in сheck_LkUsrs[kk]:
									notLikedPsts.append(kk)
							if len(notLikedPsts) == 0:
								checkLiked = True
							else:
								checkLiked = notLikedPsts

							if group_name != 'like':
								not_norm_comms = []
								notComPsts = []
								chck_avtor_com = None
								if grps_comms[allUsers.index(i)] == " ":
									chck_avtor_com = i
								else:
									chck_avtor_com = grps_comms[allUsers.index(i)]
								for kk in check_CmUsrs:
									if int(str(chck_avtor_com).replace('@club', '-')) not in check_CmUsrs[kk]:
										if int(i) not in check_CmUsrs[kk]:
											notComPsts.append(kk )
										else:
											if group_name != 'standart':
												if comment_is_normal(int(i), check_CmUsrs[kk][int(i)]) == False:
													not_norm_comms.append(kk)
									else:
										if group_name != 'standart':
											if comment_is_normal(int(i), check_CmUsrs[kk][int(str(chck_avtor_com).replace('@club', '-'))]) == False:
												not_norm_comms.append(kk)
								if len(notComPsts) == 0:
									CheckCmts = True
								else:
									CheckCmts = notComPsts
							else:
								CheckCmts = True


							# Запись данных в Excel опираясь на результаты проверки
							if checkLiked == True:
								checkLiked = []
							if CheckCmts == True:
								CheckCmts = []

							cell = ws['D'+str(n)]

							if len(checkLiked) == 0 and len(CheckCmts) == 0:
								cell.fill = PatternFill(fill_type='solid', fgColor='01FF01')
							elif len(checkLiked) != len(posts) or len(CheckCmts) != len(posts):

								if len(checkLiked) > 10 or len(CheckCmts) > 7:
									cell.fill = PatternFill(fill_type='solid', fgColor='FF0000')
								else:
									msg = ''
									if len(checkLiked) != 0:
										msg += 'Не пролайкал: '
										for k in checkLiked:
											msg += str(posts.index(k)+1)+', '
											# msg += k+', '
									if len(CheckCmts) != 0:
										if msg != '':
											msg += ' / '
										msg += 'Не прокомментировал: '
										for k in CheckCmts:
											if k in not_norm_comms:
												msg += str(posts.index(k)+1)+'*, '
											else:
												msg += str(posts.index(k)+1)+', '
											# msg += k+', '
									cell.value = msg
									cell.fill = PatternFill(fill_type='solid', fgColor='FFBC01')

							else:
								cell.fill = PatternFill(fill_type='solid', fgColor='FF0000')

						cell = ws['E2']
						cell.value = datetime.now().strftime('%d.%m.%y')
						wb.save(file_name)

					wb = openpyxl.load_workbook(file_name)
					ws = wb['Sheet1']
					wb.remove(ws)
					wb.save(file_name)

					# Загрузка файла
					# upload = VkUpload(ses_gr)
					# doc = upload.document_message(file_name, title=file_name.replace('data/', ''), peer_id=2000000001)['doc']
					# print(doc)
					upload = VkUpload(session)
					doc = upload.document(file_name, title=file_name.replace('data/', ''))['doc']

					pstsNum = ''
					n = 0
					for pp in posts:
						n += 1
						pstsNum += str(n)+'. '+pp+'\n'

					vk.messages.send(
						chat_id=56,
						message = 'Проверка в группах '+ group_name +' - '+str(g_nums[0])+' и ' + str(g_nums[1]) +' завершена. Вот файл Excel:\n'+'Также вот пронумерованный список постов, для проверки заметок из файла Excel:\n\n'+pstsNum,
						attachment=f'doc{str(doc["owner_id"])}_{str(doc["id"])}',
						random_id = 0
					)

					



		if g_n == None:
			
			# отправка всех постов данному отделу и его группам
			for p_id in bot_in_chats:
				try:
					inf = vk_gr.messages.getConversationsById(peer_ids=p_id)
					chat_name = inf['items'][0]['chat_settings']['title'].lower()
					if 'easystart' in chat_name and 'группа' in chat_name and group_name in chat_name:
						g_num = chat_name.split()[0]
						g_nums = []
						g_nums.append(int(g_num))
						ev = is_even(int(g_num)) 
						if ev == True:
							g_nums.append(int(g_num)-1)
						else:
							g_nums.append(int(g_num)+1)

						pr_gr = promotion_group(group_name, g_num)
						today_date = datetime.now().strftime("%d.%m.%y")
						partners_posts = open('data/partners_posts.txt', encoding="utf-8").read() if os.path.exists('data/partners_posts.txt') else ''
						if g_nums[0] > g_nums[1]:
							td_posts = open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt", encoding="utf-8").read() if os.path.exists(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[0])}_{str(g_nums[1])}.txt") else ''
							td_posts += partners_posts
							msg = ''
							n = 0
							for ppst in td_posts.split():
								n+=1
								msg += f'{n}) {ppst}\n\n'
							if len(td_posts.split()) > 0:
								vk_gr.messages.send(peer_id=p_id, message="@all Посты "+str(pr_gr)+" группы\n\n"+msg_send_all_posts[group_name]+"\n\n"+partners_posts+msg, random_id=0)
						else:
							td_posts = open(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt", encoding="utf-8").read() if os.path.exists(f"data/{today_date}posts_archive_{group_name}_{str(g_nums[1])}_{str(g_nums[0])}.txt") else ''
							td_posts += partners_posts
							msg = ''
							n = 0
							for ppst in td_posts.split():
								n+=1
								msg += f'{n}) {ppst}\n\n'
							if len(td_posts.split()) > 0:
								vk_gr.messages.send(peer_id=p_id, message="@all Посты "+str(pr_gr)+" группы\n\n"+msg_send_all_posts[group_name]+"\n\n"+partners_posts+msg, random_id=0)
				except BaseException:
					print(traceback.format_exc())
					vk_gr.messages.send(chat_id=1, message="ошибка в отправке сообщений Check 3", random_id=0)
	if g_n == None:
		write_prom_group(True)


def are_comments_closed(post_id, owner_id):
	try:
		f = random.choice([vkSrv1,vkSrv2,vkSrv3,vkSrv4,vkSrv5,vkSrv6]).wall.getComments(owner_id=owner_id, post_id=post_id)
		if f['can_post'] == False:
			# print(f'close: https://vk.com/wall{str(owner_id)}_{str(post_id)}')
			return 'close'
		else:
			return 'false'
	except vk_api.exceptions.ApiError as e:
		if e.code == 15:
			return 'delete'
		if e.code == 212:
			return 'close'


def PostHandler(text):
	text = text.replace('.ru', '.com').replace('.me', '.com').replace('https://m.vk.com/', 'https://vk.com/').replace('https://m.vk.ru/', 'https://vk.com/').replace('https://m.vk.me/', 'https://vk.com/')
	frgm = ['?w=', '?hash=', '?access_key=']
	for i in frgm:
		text = text.split(i)[0]
	return text


def CheckLikes(posts, vkCheck):
	chunks = [posts[i:i+9] for i in range(0, len(posts), 9)]
	psts = {}
	LkUs = {}
	psts1 = []
	for c in chunks:
		code = '''var likes = [];'''
		n=0
		for p in c:
			if p not in psts1:
				n+=1
				psts[n] = p
				psts1.append(p)
				ForOwnANdItem = p.replace("https://vk.com/wall", "").split("_") 
				owner_id = ForOwnANdItem[0]
				post_id = ForOwnANdItem[1]
				code += '''
					var post_id = %s;
					var max_count = 1000;
					var offset = %s;
					var owner = %s;
					likes = likes + {"https://vk.com/wall%s_%s": "https://vk.com/wall%s_%s"};
					var count = API.likes.getList({"owner_id": owner, "item_id": post_id, "offset": offset, "count": max_count, "type": "post"}).count;
					while(offset < count) {
						likes = likes + API.likes.getList({"owner_id": owner, "item_id": post_id, "offset": offset, "count": max_count, "type": "post"}).items;
						offset = offset + max_count;
						}
				  
				''' % (post_id, 0, owner_id, owner_id,post_id, owner_id,post_id)
		code += '''return likes;'''
		result = vkCheck.execute(code=code)
		vkCheck1 = random.choice([vkSrv1,vkSrv2,vkSrv3,vkSrv4,vkSrv5,vkSrv6])
		if vkCheck1 != vkCheck:
			vkCheck = vkCheck1

		n1 = 0
		post = None
		for i in result:
			if 'https://vk.com/wall' in str(result[i]):
				n1 += 1
				post = psts[n1]
				LkUs[post] = []
			else:
				LkUs[post].append(result[i])

	# notLikedPsts = []
	# for i in LkUs:
	# 	if int(id) not in LkUs[i]:
	# 		notLikedPsts.append(i)
	# if len(notLikedPsts) == 0:
	# 	return True
	# else:
	# 	return notLikedPsts
	return LkUs


def CheckComments(posts, vkCheck):
	vk = vkCheck
	chunks = [posts[i:i+8] for i in range(0, len(posts), 8)] # Создание чанков по 8 постов в каждом
	psts = {}
	cmUsers = {}

	for c in chunks:
		script = '''var comments = [];'''
		
		n=0
		for i in c:
			n+=1
			psts[n] = i
			ForOwnANdItem = i.replace("https://vk.com/wall", "").split("_") 
			owner_id = ForOwnANdItem[0]
			post_id = ForOwnANdItem[1]

			# Шаблон скрипта для получения комментариев
			script += '''
				var post_id = %s;
				var max_count = 100;
				var offset = %s;
				 var owner = %s;
				comments = comments + {"https://vk.com/wall%s_%s": "https://vk.com/wall%s_%s"};
				var count = API.wall.getComments({"owner_id": owner, "post_id": post_id, "offset": offset, "count": max_count}).count;
				while(offset < count) {
					comments = comments + API.wall.getComments({"owner_id": owner, "post_id": post_id, "offset": offset, "count": max_count}).items;
					offset = offset + max_count;
					}
			''' % (post_id, 0, owner_id, owner_id,post_id, owner_id,post_id)

		# Отправляем запрос на получение всех комментариев
		script += '''return comments;'''
		cm = vkCheck.execute(code=script)
		n1 = 0
		
		post = None
		for m in cm:
			if 'https://vk.com/wall' in m:
				n1 += 1
				cmUsers[psts[n1]] = {}
				post = psts[n1]
			else:
				cmUsers[psts[n1]][cm[m]['from_id']] = cm[m]['text']

	return cmUsers


def is_even(number):
	return number % 2 == 0


def comment_is_normal(id, text):
	# threading.Thread(target=send_gr, args=(f"Комментарий от @id{str(id)}:\n\n"+text, )).start()
	if text == '':
		return False
	else:
		if len(text.split())<3:
			return False
		else:
			return True


def send_gr(text):
	vk_gr.messages.send(chat_id=1, message=text, random_id=0)


if __name__ == '__main__':
	threading.Thread(target=main).start()
	threading.Thread(target=timer_check).start()


	

# про говнюка была шутка, закрывай код и пиздуй, проверять меня захотел 
# (адресовано программисту который будет проверять)